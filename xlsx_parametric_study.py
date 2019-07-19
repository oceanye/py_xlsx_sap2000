#coding=UTF-8

import openpyxl
import os

path = os.getcwd()
files = os.listdir(path)
f_n ='' .join(files)


if f_n.count('.xlsx')<>1:
    print'only one xlsx file is allowed'
    quit()
else:
    for ff in files:
        if ff.find('.xlsx')>0:
            file_name=ff
            break
print file_name

workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.get_sheet_by_name("Frame Loads - Distributed")

#Frame Loads - Distributed

[r,c]=['4','J']

if len(c)==1:
    ind_c = ord(c)-64
else:
    ind_c = (ord(c[0])-64)*26+(ord(c[1])-64)

ind_r = int(r)

print ind_r,ind_c

par = range(1,10,1)
par_label = 'load_mod'
for p in par:
   #content_A1 = worksheet.cell(row=ind_r,column=ind_c).value
   #print content_A1
    worksheet.cell(ind_r,ind_c,p)
    p_file_name = file_name.replace('.xlsx','_'+par_label+'_'+str(p)+'.xlsx')
    workbook.save(filename=p_file_name)

